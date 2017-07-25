"""
Name: Michael Katcher
Date: 07/24/2017
Desc: A script to solve the following problem:

    Background:     A folder with an .xlsm for each trading day
                    In each file is a data tab with the same name

    Requirement:    Collect all the tabs for a given date range
                    and save them in a single workbook with one
                    tab for each date.
"""


from os import listdir
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook


def date_from_file_name(file_name, return_datetime=False):
    """
    Return full date as a string or datetime from a date in a filename.

    Args:
        file_name (str): Formatted as: Trade E-mail Archive - yyyy.mm.dd.xlsm
        return_datetime (bool, optional): TRUE if return should be datetime
                                      FALSE if return should be string

    Returns:
        str or datetime: The return value. Date from filename as yyyy.mm.dd
    """
    if return_datetime:
        year = int(file_name[-15:-11])
        month = int(file_name[-10:-8])
        day = int(file_name[-7:-5])
        return datetime(year, month, day)
    else:
        return file_name[-15:-5]


def copy_worksheet(ws_input, ws_output):
    """
    Copy values from one openpyxl Worksheet to another.

     Args:
         ws_input (Worksheet): The worksheet the data is being copied from.
         ws_output (Worksheet): The worksheet the data is being copied to.
    """
    for row in ws_input.rows:
        ws_output.append([cell.value for cell in row])


class OutputWorkbook(Workbook):
    """An extension of the Workbook class to add a context manager."""
    def __init__(self, filename):
        """Extends Workbook.__init__ to add a filename and defaults to write_only.

        Args:
            filename (str): The path and filename where the output workbook is saved.
        """
        self._filename = filename

        super(OutputWorkbook,self).__init__(write_only=True)

    def __enter__(self):
        return self

    def __exit__(self, *args):
        print 'Saving %s... ' % self._filename.split('\\')[-1]
        self.save(filename=self._filename)
        print 'Save Complete!'


def main():
    # Initial setup variables avoid excessive blank lines
    data_folder = 'N:\\~~~\\Trade Email\\' # Folder containing the input files
    start_date = datetime(2017,07,01) # The date of the first file you want to import
    end_date = datetime.now() # The date of the last file you want to import
    sheet_name = 'Data' # The sheet name we're looking for in each inpupt file
    save_path = 'H:\\Test.xlsx' # The path of the output file

    # Generate list of filenames
    def include_file(name): return name.endswith('.xlsm') and start_date <= date_from_file_name(name,True) <= end_date
    file_names = [name for name in listdir(data_folder) if include_file(name)]

    with OutputWorkbook(save_path) as wb_output:
        for idx,file_name in enumerate(file_names):
            print 'Processing File %s of %s' % (idx + 1, len(file_names))

            ws_output = wb_output.create_sheet(date_from_file_name(file_name))
            ws_input = load_workbook(data_folder + file_name, read_only=True)[sheet_name]

            copy_worksheet(ws_input,ws_output)


if __name__ == '__main__':
    main()
